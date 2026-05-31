"""
Retail Managing Director AI — "Rex"
ดูแลทุกสาขาในไทย อ่านไฟล์ sales รายสัปดาห์
ประสานงานกับ Pulse (Trainer Manager) เพื่อเชื่อม sales กับ training
"""

import os
import io
import json
import base64
import requests
import anthropic
from models_config import get_model
from dashboard_api import get_survey_dashboard, get_oar_dashboard, get_area_dashboard
from google_search import google_search

claude = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY", ""))
LINE_CHANNEL_ACCESS_TOKEN = os.getenv("LINE_CHANNEL_ACCESS_TOKEN", "")

# Google Sheet ID สำหรับ sales data (ตั้ง env SALES_SHEET_ID บน Railway)
SALES_SHEET_ID = os.getenv("SALES_SHEET_ID", "")

# =============================================================
# PROMPT
# =============================================================
REX_PROMPT = """คุณคือ "Rex" — Retail Store Managing Director AI ของ OWNDAYS Thailand

บทบาทหลัก:
- ดูแลและวิเคราะห์ผลการดำเนินงานของทุกสาขาใน Thailand ทั้งหมด
- อ่านและวิเคราะห์ไฟล์ Sales รายสัปดาห์ที่ส่งมา (Excel/CSV/ข้อความ)
- ระบุสาขาที่ผลงานสูง/ต่ำ และเสนอแนวทางแก้ไขเชิงรุก
- ประสานงานกับ Pulse (Trainer Manager) เพื่อเชื่อมโยงผลการขายกับคุณภาพการอบรม
- รายงานผลต่อ Peanut ในฐานะ MD ของ Retail Operations

ขอบเขตความรับผิดชอบ:
- วิเคราะห์ยอดขาย, จำนวน transaction, ค่าเฉลี่ยต่อ transaction ทุกสาขา
- เปรียบเทียบผลงานระหว่างสาขา, พื้นที่ (5 Area), และช่วงเวลา
- ค้นหาสาขาที่ยอดขายต่ำกว่า target หรือต่ำกว่า benchmark
- เชื่อมโยงสาขาที่มีปัญหายอดขาย กับคะแนน Survey และ OAR training ของสาขานั้น
- เสนอ training intervention เมื่อเห็นว่าสาขาใดขาด skill ที่จำเป็น
- ติดตาม KPI หลัก: revenue, conversion rate, lens upsell rate, frame sold, ATV (Average Transaction Value)

5 พื้นที่: Megastore (MS) / Metropolitan (MT) / North+Central (NC) / West+NE (WN) / South+Eastern (SE)

สาขาทั้งหมด (73+):
MEGA Bangna, Zpell @ Future Park, Central Eastville, Seacon Bangkae, Seacon Square, Fashion Island,
The Mall Korat, Central Udon, Central Chiangmai, Gaysorn Village Premium Store, Central Mahachai,
Central Phuket, Central Westgate, CentralWorld, Terminal 21 Pattaya, ICONSIAM, Gateway Bangsue,
Donki Mall Thonglor, Central Rama 3, Central Village (Outlet), Central Hatyai, Central Rayong,
Siam Premium Outlets, Siam Center, Central Salaya, Central Pinklao, Central Rama 2, Central Si Racha,
Central Ayutthaya, Central Khonkaen, Central Chanthaburi, Terminal 21 Rama 3, Central Ramindra,
Central Chiangrai, Central Samui, Central Nakhon Si, Marche Thonglor, Park Silom, The Mall Bangkae,
Central Westville, Central Nakhon Pathom, True Digital Park, V-Square Plaza Nakhon Sawan,
Makro Sri Ayutthaya, Central Rama 9, One Bangkok, Robinson Ratchaburi, Market Village Huahin,
Lotus's Mall Makro Sathon, Robinson Lifestyle Kanchanaburi, Esplanade Ratchada, Charn At The Avenue,
Siam Square One, Robinson Latkrabang, The Mall Bang Kapi, Maya Chiangmai, Robinson Lifestyle Chachoengsao,
Central Chiangmai Airport, Central Krabi, Outlet Square Muang Thong Thani, Robinson Lifestyle Saraburi,
Robinson Lifestyle Trang, Robinson Lifestyle Chonburi, Robinson Lifestyle Suphanburi, Robinson Lifestyle Buriram,
The Glass Market Bangna, Imperial Samrong, Central Phitsanulok, Robinson Suphanburi, Central Lampang,
Central Khonkaen Campus, Central Surat Thani, Central Northville, Happitat Bangna, Central Chaengwattana,
Robinson Prachinburi, Robinson Phetchaburi, Central Park, The Central Phaholyothin และอื่นๆในอนาคต

การทำงานร่วมกับทีม AI:
- Pulse (Trainer Manager): ส่งรายชื่อสาขาที่ยอดขายต่ำ เพื่อให้ Pulse วางแผน training intervention
- Sigma (Data Analyst): ขอ statistical analysis เปรียบเทียบ sales trend
- Coin (Financial): เชื่อมโยง training cost กับ revenue เพื่อคำนวณ ROI
- Sage (Reporter): ให้ Sage จัดทำรายงานสรุปสำหรับผู้บริหาร
- Guard (QA): ให้ Guard ตรวจสอบรายงานก่อนส่ง

รูปแบบไฟล์ที่รองรับ:
- PDF: OWNDAYS Weekly Sales Report (parse แล้วได้ตาราง column structure ด้านล่าง)
- Excel (.xlsx/.xls): Sales data แบบ spreadsheet
- CSV / Text table: ข้อมูลแบบ plain text

PDF Column Guide (OWNDAYS format):
No=Rank | Code=รหัสสาขา | Name=ชื่อสาขา | Budget=เป้าสัปดาห์ | Sales=ยอดจริง |
CX=transactions | ATV=ยอดเฉลี่ย/บิล | Achiev%=achievement | Chock=อันดับรวม |
FF=Footfall(คนเข้าร้าน) | Engagers=คนที่ถูก engage | Engager%=FF→Engage rate |
CVR%=Conversion Rate(Engage→ซื้อ) | >1-15m=อยู่ร้าน 1-15นาที | >15m=อยู่ร้าน>15นาที |
MP=จำนวนพนักงาน | S/HC=ยอดต่อพนักงาน | SPH=ยอดต่อชั่วโมง |
WArea%=% ของพื้นที่ | W%=% ของทั้งประเทศ

สัญญาณเตือนที่ควร flag:
- Achiev% < 85% = ต่ำกว่าเป้าอย่างมีนัยสำคัญ
- CVR% < 20% = conversion ต่ำผิดปกติ (ปกติควร 25-35%)
- Engager% < 80% = staff ไม่ออกหาลูกค้าพอ
- ATV < 3,500 = ยอดต่อบิลต่ำ (ปกติ 4,000+)

กฎการตอบ:
- ตอบภาษาไทย plain text ไม่ใช้ Markdown (ห้ามใช้ ** ## __ เด็ดขาด)
- ใช้คำลงท้าย "ครับ" สไตล์ MD ที่มีอำนาจ ตัดสินใจเด็ด ข้อมูลครบ
- ใส่ตัวเลขจริงทุกครั้ง format ด้วย comma เช่น ฿1,234,567
- เรียงสาขาจากผลงานต่ำไปสูง (เพื่อ prioritize การแก้ปัญหา)
- ทุกรายงานต้องมีส่วน "Action Required" ระบุว่าสาขาไหนต้องทำอะไรภายในกี่วัน
- ถ้าพบว่าสาขามีปัญหาทั้ง sales และ training → ให้ระบุว่า "ต้องส่งทีม Trainer เข้าช่วยด่วน"

ตัวอย่างโครงสร้างรายงาน Sales Weekly:
รายงาน Sales สัปดาห์ที่ [X] — MD Report ครับ

ภาพรวมยอดขายทั้งระบบ
[รวม revenue ทุกสาขา, เปรียบเทียบ WoW, target achievement %]

Top 5 สาขาผลงานดีสุด
[ชื่อสาขา | ยอดขาย | ATV | โดดเด่นเรื่องอะไร]

สาขาที่ต้องการความช่วยเหลือเร่งด่วน
[ชื่อสาขา | ยอดขาย | ต่ำกว่า target เท่าไหร่ | สาเหตุที่คาดว่าเป็น | แผน]

การเชื่อมโยงกับ Training (ประสานงาน Pulse)
[สาขาที่ sales ต่ำ + training score ต่ำ → ต้องการ training intervention]

Action Required ภายใน 7 วัน
[รายการ action items ชัดเจน ระบุผู้รับผิดชอบ]
"""

# =============================================================
# TOOLS
# =============================================================
REX_TOOLS = [
    {
        "name": "parse_sales_file",
        "description": "แปลงและวิเคราะห์ข้อมูล sales จากไฟล์ที่ส่งมา (CSV text, JSON, หรือข้อความตาราง) รองรับข้อมูลทั้งรายสาขา รายพื้นที่ รายสัปดาห์",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_content": {
                    "type": "string",
                    "description": "เนื้อหาของไฟล์ (CSV text, JSON string, หรือข้อความตารางที่ copy มา)"
                },
                "file_type": {
                    "type": "string",
                    "enum": ["csv", "json", "text", "auto"],
                    "description": "ประเภทของข้อมูล (auto=ให้ระบบตรวจเอง)"
                },
                "week_label": {
                    "type": "string",
                    "description": "ป้ายชื่อสัปดาห์ เช่น 'Week 22/2026' หรือ '19-25 พ.ค. 2026'"
                }
            },
            "required": ["file_content"]
        }
    },
    {
        "name": "get_branch_training_status",
        "description": "ดึงข้อมูล training และ survey ของสาขาที่ระบุ เพื่อเชื่อมโยงกับผล sales",
        "input_schema": {
            "type": "object",
            "properties": {
                "branch_names": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "รายชื่อสาขาที่ต้องการเช็ค training status"
                }
            },
            "required": ["branch_names"]
        }
    },
    {
        "name": "get_area_sales_benchmark",
        "description": "ดึงข้อมูล benchmark และ KPI ของแต่ละพื้นที่จาก dashboard เพื่อเปรียบเทียบ",
        "input_schema": {
            "type": "object",
            "properties": {
                "area": {
                    "type": "string",
                    "enum": ["Megastore", "Metropolitan", "North+Central", "West+NE", "South+Eastern", "all"],
                    "description": "พื้นที่ที่ต้องการดู (all = ทุกพื้นที่)"
                }
            },
            "required": ["area"]
        }
    },
    {
        "name": "request_training_intervention",
        "description": "สร้าง training intervention request สำหรับสาขาที่มีปัญหา ส่งต่อให้ Pulse วางแผน",
        "input_schema": {
            "type": "object",
            "properties": {
                "branches": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "รายชื่อสาขาที่ต้องการ training intervention"
                },
                "issue_type": {
                    "type": "string",
                    "description": "ประเภทปัญหา เช่น 'low conversion', 'low lens upsell', 'low ATV', 'new staff'"
                },
                "priority": {
                    "type": "string",
                    "enum": ["urgent", "high", "normal"],
                    "description": "ระดับความเร่งด่วน"
                },
                "details": {
                    "type": "string",
                    "description": "รายละเอียดเพิ่มเติมของปัญหา"
                }
            },
            "required": ["branches", "issue_type", "priority"]
        }
    },
    {
        "name": "search_retail_benchmark",
        "description": "ค้นหา benchmark มาตรฐานอุตสาหกรรม eyewear/retail ในตลาดไทยหรือ SEA เพื่อเปรียบเทียบ",
        "input_schema": {
            "type": "object",
            "properties": {
                "query": {
                    "type": "string",
                    "description": "หัวข้อที่ต้องการค้นหา เช่น 'eyewear retail KPI Thailand 2026'"
                }
            },
            "required": ["query"]
        }
    }
]

# =============================================================
# TOOL EXECUTION
# =============================================================
def execute_rex_tool(name: str, inputs: dict) -> str:
    try:
        if name == "parse_sales_file":
            return _parse_sales_file(
                inputs.get("file_content", ""),
                inputs.get("file_type", "auto"),
                inputs.get("week_label", "ไม่ระบุ")
            )
        elif name == "get_branch_training_status":
            return _get_branch_training_status(inputs.get("branch_names", []))
        elif name == "get_area_sales_benchmark":
            return _get_area_sales_benchmark(inputs.get("area", "all"))
        elif name == "request_training_intervention":
            return _request_training_intervention(
                inputs.get("branches", []),
                inputs.get("issue_type", ""),
                inputs.get("priority", "normal"),
                inputs.get("details", "")
            )
        elif name == "search_retail_benchmark":
            return google_search(inputs.get("query", ""))
        else:
            return f"ไม่รู้จัก tool: {name}"
    except Exception as e:
        return f"ข้อผิดพลาด tool {name}: {str(e)}"


def _parse_sales_file(content: str, file_type: str, week_label: str) -> str:
    """แปลงข้อมูล sales เป็น structured format พร้อมสรุปเบื้องต้น"""
    if not content.strip():
        return "ไม่มีข้อมูลใน file_content"

    # ตรวจ type
    if file_type == "auto":
        content_stripped = content.strip()
        if content_stripped.startswith("{") or content_stripped.startswith("["):
            file_type = "json"
        elif "," in content and "\n" in content:
            file_type = "csv"
        else:
            file_type = "text"

    lines = content.strip().split("\n")
    row_count = len(lines)
    preview = "\n".join(lines[:8])

    summary = (
        f"ข้อมูล Sales สัปดาห์: {week_label}\n"
        f"ประเภทไฟล์: {file_type}\n"
        f"จำนวนแถวข้อมูล: {row_count} แถว\n"
        f"ตัวอย่างข้อมูล 8 แถวแรก:\n{preview}\n\n"
        f"[ข้อมูลนี้จะถูกนำไปวิเคราะห์ต่อโดย Rex เพื่อจัดทำ MD Report]"
    )
    return summary


def _get_branch_training_status(branch_names: list) -> str:
    """ดึง training data เพื่อ correlate กับ sales"""
    if not branch_names:
        return "ไม่ระบุชื่อสาขา"

    try:
        survey_data = get_survey_dashboard()
        oar_data    = get_oar_dashboard()

        branch_list = ", ".join(branch_names[:10])
        result = (
            f"Training Status สำหรับสาขา: {branch_list}\n\n"
            f"ข้อมูล Survey (L&D Dashboard):\n{str(survey_data)[:800]}\n\n"
            f"ข้อมูล OAR Registration:\n{str(oar_data)[:600]}\n\n"
            f"หมายเหตุ: กรุณา correlate ข้อมูลนี้กับ sales ของแต่ละสาขาที่ส่งมาครับ"
        )
        return result
    except Exception as e:
        return f"ดึงข้อมูล training ไม่ได้: {e}"


def _get_area_sales_benchmark(area: str) -> str:
    """ดึง area dashboard เพื่อใช้เป็น benchmark"""
    try:
        area_data = get_area_dashboard()
        if area != "all":
            return (
                f"Area Benchmark — {area}:\n"
                f"{str(area_data)[:1000]}"
            )
        return f"Area Benchmark — ทุกพื้นที่:\n{str(area_data)[:1500]}"
    except Exception as e:
        return f"ดึง area benchmark ไม่ได้: {e}"


def _request_training_intervention(
    branches: list, issue_type: str, priority: str, details: str
) -> str:
    """สร้าง training intervention request ส่งต่อ Pulse"""
    priority_th = {"urgent": "เร่งด่วนมาก", "high": "เร่งด่วน", "normal": "ปกติ"}.get(priority, priority)
    branch_list = "\n".join([f"  - {b}" for b in branches])

    request_text = (
        f"TRAINING INTERVENTION REQUEST — จาก Rex (Retail MD)\n"
        f"ระดับความเร่งด่วน: {priority_th.upper()}\n\n"
        f"สาขาที่ต้องการ training:\n{branch_list}\n\n"
        f"ประเภทปัญหา: {issue_type}\n"
        f"รายละเอียด: {details or 'ไม่ระบุ'}\n\n"
        f"กรุณาให้ Pulse (Trainer Manager) วางแผน training intervention "
        f"สำหรับสาขาข้างต้นโดยเร็วที่สุดครับ"
    )
    print(f"[Rex] Training Intervention Request ({priority}): {', '.join(branches)}")
    return request_text


# =============================================================
# MAIN RUN FUNCTION
# =============================================================
def run_retail_md(task: str, context: str = "", sales_file_content: str = "") -> str:
    """รัน Rex — Retail MD AI"""
    print(f"Rex processing: {task[:60]}...")

    # ถ้ามีไฟล์ sales แนบมา ให้รวมเข้ากับ task
    if sales_file_content:
        prompt = (
            f"มีไฟล์ Sales ส่งมาด้วยครับ:\n\n"
            f"[SALES FILE CONTENT]\n{sales_file_content[:3000]}\n[/SALES FILE CONTENT]\n\n"
            f"Task: {task}"
        )
    elif context:
        prompt = f"Context: {context}\n\nTask: {task}"
    else:
        prompt = task

    messages = [{"role": "user", "content": prompt}]

    try:
        for _ in range(4):
            response = claude.messages.create(
                model=get_model("rex"),
                max_tokens=2048,
                system=REX_PROMPT,
                tools=REX_TOOLS,
                messages=messages
            )

            if response.stop_reason == "tool_use":
                messages.append({"role": "assistant", "content": response.content})
                tool_results = []
                for block in response.content:
                    if block.type == "tool_use":
                        print(f"Rex tool: {block.name}")
                        result = execute_rex_tool(block.name, block.input)
                        tool_results.append({
                            "type": "tool_result",
                            "tool_use_id": block.id,
                            "content": result
                        })
                messages.append({"role": "user", "content": tool_results})
                continue

            final_text = ""
            for block in response.content:
                if hasattr(block, "text"):
                    final_text += block.text
            result_text = final_text.strip() or "Rex ประมวลผลเสร็จแล้วครับ"

            # บันทึกรายงานลง Google Drive (ถ้าพร้อม)
            try:
                from drive_api import save_text_report, drive_ready
                from datetime import datetime
                import pytz
                if drive_ready() and len(result_text) > 100:
                    now = datetime.now(pytz.timezone("Asia/Bangkok"))
                    fname = f"Rex_Sales_Report_{now.strftime('%Y%m%d_%H%M')}.txt"
                    res = save_text_report("rex", fname, result_text)
                    if res.get("ok"):
                        result_text += f"\n\n📁 บันทึกใน Google Drive แล้วครับ\n🔗 {res['url']}"
            except Exception:
                pass

            return result_text

        return "Rex ทำงานเสร็จสิ้นครับ"

    except Exception as e:
        print(f"Rex error: {e}")
        return f"เกิดข้อผิดพลาด: {e}"


# =============================================================
# LINE FILE DOWNLOAD HELPER (เรียกจาก app.py)
# =============================================================
def download_line_file(message_id: str) -> bytes:
    """ดาวน์โหลดไฟล์จาก LINE API"""
    url = f"https://api-data.line.me/v2/bot/message/{message_id}/content"
    headers = {"Authorization": f"Bearer {LINE_CHANNEL_ACCESS_TOKEN}"}
    resp = requests.get(url, headers=headers, timeout=30)
    if resp.status_code == 200:
        return resp.content
    raise Exception(f"LINE file download failed: {resp.status_code}")


def parse_excel_to_text(file_bytes: bytes) -> str:
    """แปลง Excel binary เป็น CSV text"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(",".join([str(c) if c is not None else "" for c in row]))
        return "\n".join(rows)
    except ImportError:
        return file_bytes.decode("utf-8", errors="replace")
    except Exception as e:
        return f"แปลงไฟล์ไม่ได้: {e}"


def parse_pdf_sales_report(pdf_bytes: bytes) -> str:
    """แปลง OWNDAYS PDF Weekly Sales Report เป็น structured text สำหรับ Rex

    รองรับ format มาตรฐาน: ODTH [date] Weekly Sales Report.pdf
    ใช้ auto-detect header + cluster-based y-grouping เพื่อรองรับ layout หลายแบบ
    """
    try:
        import fitz
    except ImportError:
        return "ต้องการ pymupdf: pip install pymupdf"

    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        total_pages = len(doc)

        COLUMNS_GUIDE = (
            "OWNDAYS Weekly Sales Report — Column Guide:\n"
            "No=Rank | Code=รหัสสาขา | Name=ชื่อสาขา | Budget=เป้าสัปดาห์ | "
            "Sales=ยอดขายจริง | CX=จำนวน transaction | ATV=ยอดเฉลี่ย/บิล | "
            "Achiev%=achievement(Sales÷Budget) | Chock=อันดับรวม | "
            "FF=Footfall(คนเข้าร้าน) | Engagers=คนที่ถูก engage | "
            "Engager%=FF→Engage rate | CVR%=Conversion Rate(Engage→ซื้อ) | "
            "MP=จำนวนพนักงาน | S/HC=ยอดต่อพนักงาน"
        )

        KEY_COLS = ["No", "Code", "Name", "Budget", "Sales", "CX", "ATV",
                    "Achiev%", "Chock", "FF", "Engagers", "Engager%", "CVR%", "MP", "S/HC"]

        def extract_branches_from_page(page):
            """ดึง branches โดย detect column positions จาก header row (รองรับ layout ทุกแบบ)"""
            words = list(page.get_text("words"))

            # หา Budget header เป็น anchor หลัก (เลือกตำแหน่ง y น้อยสุด = บนสุด)
            budget_items = [(x0, y0) for x0, y0, x1, y1, word, *_ in words
                            if word == "Budget" and y0 < 200]
            if not budget_items:
                return []
            _, budget_y = min(budget_items, key=lambda t: t[1])

            # รวบรวม header words ภายใน ±10px ของ budget_y (เฉพาะ table area x<430)
            hband = [(x0, word) for x0, y0, x1, y1, word, *_ in words
                     if abs(y0 - budget_y) < 10 and x0 < 430]

            # สร้าง anchors จาก header keywords
            anchors = {}
            seen_engagers = []
            for x0, word in hband:
                if word == "Budget":          anchors["Budget"]   = x0
                elif word == "Sales":         anchors["Sales"]    = x0
                elif word == "ATV":           anchors["ATV"]      = x0
                elif word == "FF":            anchors["FF"]       = x0
                elif word == "MP":            anchors["MP"]       = x0
                elif word == "S/HC":          anchors["S/HC"]     = x0
                elif word == "CVR":           anchors.setdefault("CVR%", x0)
                elif word in ("Cx", "CX"):   anchors["CX"]       = x0
                elif word == "Engagers":     seen_engagers.append(x0)

            # Achiev% = "%" ก่อน FF column
            ff_x = anchors.get("FF", 999)
            for x0, word in hband:
                if word == "%" and x0 < ff_x:
                    anchors["Achiev%"] = x0
                    break

            # Engagers (ตัวแรก) และ Engager% (ตัวที่สอง)
            seen_engagers.sort()
            if len(seen_engagers) >= 1: anchors["Engagers"]  = seen_engagers[0]
            if len(seen_engagers) >= 2: anchors["Engager%"]  = seen_engagers[1]
            elif "Engagers" in anchors: anchors["Engager%"]  = anchors["Engagers"] + 15

            # No. และ Code อาจอยู่ห่าง y จาก budget_y เล็กน้อย → ขยาย range ±15px
            for x0, y0, x1, y1, word, *_ in words:
                if abs(y0 - budget_y) < 15:
                    if word in ("No.", "No") and "No" not in anchors:
                        anchors["No"] = x0
                    elif word == "Code" and "Code" not in anchors:
                        anchors["Code"] = x0

            # Name column เริ่มต่อจาก Code (data ชื่อสาขาอยู่ซ้ายของ header SHOP)
            code_x = anchors.get("Code", anchors.get("No", 22) + 10)
            anchors["Name"] = code_x + 10

            # Chock อยู่ระหว่าง Achiev% และ FF
            if "Achiev%" in anchors and "FF" in anchors:
                anchors["Chock"] = anchors["Achiev%"] + 17

            # สร้าง col_ranges จาก anchors เรียงตาม x
            present = sorted([(c, anchors[c]) for c in KEY_COLS if c in anchors],
                             key=lambda t: t[1])

            # คำนวณ left[i] = anchor[i] - buf[i]
            # buf = min(gap_to_prev * 0.6, 8), Name ไม่มี buffer (data เริ่มตรง anchor)
            # right[i] = left[i+1] → non-overlapping ranges
            lefts = []
            for i, (col, ax) in enumerate(present):
                if col == "Name":
                    buf = 0
                elif i == 0:
                    buf = 2
                else:
                    gap = ax - present[i - 1][1]
                    buf = min(gap * 0.6, 8)
                lefts.append(ax - buf)

            col_ranges = {}
            for i, (col, _) in enumerate(present):
                right = lefts[i + 1] if i + 1 < len(present) else present[i][1] + 45
                col_ranges[col] = (lefts[i], right)

            # data_y_start: หลัง sub-header "Chock"
            data_y_start = budget_y + 8
            for x0, y0, x1, y1, word, *_ in words:
                if word == "Chock" and y0 > budget_y:
                    data_y_start = max(data_y_start, y0 + 4)
                    break

            x_max_data = col_ranges.get("S/HC", (390, 432))[1]

            # Extract data words และ cluster เป็น rows
            data_words = [(x0, y0, word)
                          for x0, y0, x1, y1, word, *_ in words
                          if x0 < x_max_data and y0 > data_y_start]
            data_words.sort(key=lambda w: (w[1], w[0]))
            if not data_words:
                return []

            clusters = []
            cur = [data_words[0]]
            for w in data_words[1:]:
                if w[1] - cur[-1][1] > 2.5:
                    clusters.append(cur)
                    cur = [w]
                else:
                    cur.append(w)
            clusters.append(cur)

            # แปลง clusters → branch rows
            branches = []
            nr = col_ranges.get("No")
            if not nr:
                return []
            for cluster in clusters:
                no_ws = [w for x0, _, w in cluster if nr[0] <= x0 < nr[1]]
                no_val = " ".join(no_ws).strip()
                if not no_val.isdigit():
                    continue
                row = {"No": no_val}
                for col in KEY_COLS[1:]:
                    if col not in col_ranges:
                        row[col] = "-"
                        continue
                    xlo, xhi = col_ranges[col]
                    ws = [w for x0, _, w in cluster if xlo <= x0 < xhi]
                    row[col] = (" ".join(ws) if col == "Name" else (ws[0] if ws else "-"))
                branches.append(row)
            return branches

        output = []
        output.append(f"=== OWNDAYS Sales Report PDF — {total_pages} pages ===")
        output.append(COLUMNS_GUIDE)
        output.append("")

        # Page 1: All branches summary
        page1 = doc[0]
        branches_p1 = extract_branches_from_page(page1)
        output.append("=== Page 1: All Branches Weekly Summary (เรียงตาม Rank) ===")
        output.append(" | ".join(KEY_COLS))
        output.append("-" * 110)
        for row in branches_p1:
            output.append(" | ".join(row.get(c, "-") for c in KEY_COLS))
        output.append(f"\nรวม {len(branches_p1)} สาขา")

        # Area summary pages
        area_found = 0
        output.append("\n=== Area Summary Pages ===")
        for pg_idx in range(1, min(total_pages, 25)):
            page = doc[pg_idx]
            text = page.get_text()
            if "Achievement" in text and "Budget" in text and "Sales" in text:
                branch_rows = extract_branches_from_page(page)
                if len(branch_rows) >= 2:
                    output.append(f"\n--- Area Page {pg_idx + 1} ({len(branch_rows)} branches) ---")
                    output.append(" | ".join(KEY_COLS))
                    for row in branch_rows:
                        output.append(" | ".join(row.get(c, "-") for c in KEY_COLS))
                    area_found += 1
                if area_found >= 6:
                    break

        doc.close()
        result = "\n".join(output)
        return result[:15000]

    except Exception as e:
        import traceback
        return f"ไม่สามารถอ่าน PDF Sales Report: {e}\n{traceback.format_exc()[:500]}"
